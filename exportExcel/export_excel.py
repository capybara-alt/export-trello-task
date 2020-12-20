import trelloAPI.trello_api as trello_api
import trelloAPI.card as trello_card
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Border, Side
from datetime import datetime
import numpy as np
import calendar
import json
import csv
import os

ROW_START=1
DATA_ROW_START=3
NORMAL_SIDE = Side(style='thin', color='000000')
BORDER = Border(top=NORMAL_SIDE,bottom=NORMAL_SIDE,right=NORMAL_SIDE,left=NORMAL_SIDE)
CONVEX_BORDER = Border(top=NORMAL_SIDE,right=NORMAL_SIDE,left=NORMAL_SIDE)
CONVEX_DOWNWARD_BORDER = Border(bottom=NORMAL_SIDE,right=NORMAL_SIDE,left=NORMAL_SIDE)
TOP_FILL = PatternFill(fill_type='solid', fgColor='55FF55')
PHASE_FILL = PatternFill(fill_type='solid', fgColor='FFFF55')
DATE_FILL = PatternFill(fill_type='solid', fgColor='0000FF')
ACTUAL_DATE_FILL = PatternFill(fill_type='solid', fgColor='FF0000')

class ExportExcel:

    def __init__(self, json_file):
        self.__json_file = json_file
        self.__boards = []
        self.__wb = Workbook()
        self.__ws = self.__wb.active
        self.__ws.title = 'タスク一覧'

    def import_from_files(self):
        files = os.listdir('./jsons')
        if files.count == 0 :
            return

        for f in files:
            print('load file : ' + f)
            file_open = open('./jsons/' + f, 'r')
            json_str = json.load(file_open)
            api = trello_api.TrelloAPI(json_str)
            self.__boards.append(api.get_board())
            api.sort_cards_by_date()

    def __set_item_name_cell(self, col_num, width, name):
        if width > 0:
            self.__ws.column_dimensions[get_column_letter(col_num)].width = width
        self.__ws.cell(ROW_START, col_num).value = name
        self.__ws.cell(row=ROW_START, column=col_num).fill = TOP_FILL
        self.__ws.cell(row=ROW_START+1, column=col_num).fill = TOP_FILL
        self.__ws.cell(row=ROW_START, column=col_num).border = CONVEX_BORDER
        self.__ws.cell(row=ROW_START+1, column=col_num).border = CONVEX_DOWNWARD_BORDER

    def __set_performance_date_cell(self, max_col):
        self.__set_item_name_cell(max_col + 1, 0, '実績')

        project_start_date = trello_api.get_project_start_date(self.__boards)
        dates = []
        for i in range(0, 3):
            year = project_start_date.year
            month = project_start_date.month + i
            try:
                dates_sum = sum(calendar.Calendar().monthdatescalendar(year, month), [])
                filtered_dates = list(filter(lambda x: x.month == month, dates_sum))
                dates.append(filtered_dates)
            except calendar.IllegalMonthError:
                month = month - 12
                year += 1
                dates_sum = sum(calendar.Calendar().monthdatescalendar(year, month), [])
                filtered_dates = list(filter(lambda x: x.month == month, dates_sum))
                dates.append(filtered_dates)
        
        col_offset = self.__ws.max_column
        for date in dates:
            self.__ws.cell(ROW_START+1, col_offset).value = date[0].month
            for i in range(0, len(date)):
                self.__ws.cell(ROW_START, col_offset + i).fill = TOP_FILL
                self.__ws.cell(ROW_START+1, col_offset + i).fill = TOP_FILL
                self.__ws.cell(ROW_START, col_offset + i).border = BORDER 
                self.__ws.cell(ROW_START+1, col_offset + i).border = BORDER 
                self.__ws.cell(DATA_ROW_START, col_offset + i).fill = PHASE_FILL
                self.__ws.cell(DATA_ROW_START, col_offset + i).border = BORDER 

                self.__ws.cell(DATA_ROW_START, col_offset + i).value = date[i].day
            col_offset += len(date)
        return dates

    def __fill_task_date_in_date(self, dates, card, row, col):
        for i in range(0, len(dates)):
            for j in range(0, len(dates[i])):
                if trello_api.is_task_date_in_date(card, dates[i][j]):
                    self.__ws.cell(row, col + i * len(dates[i]) + j).fill = DATE_FILL


    def performance(self, max_col):
        dates = self.__set_performance_date_cell(max_col)
        row = DATA_ROW_START
        col = max_col + 1
        for board in self.__boards:
            row += 1
            for card in board.get_cards():
                self.__fill_task_date_in_date(dates, card, row, col)
                row += 1
                

    def task_ids(self, col_num): 
        self.__set_item_name_cell(col_num, 30, 'ID')

        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).value = board.get_board_id()
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).value = card.get_card_id() 
                row += 1

    def task_names(self, col_num):
        self.__set_item_name_cell(col_num, 30, 'タスク名')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).value = board.get_name()
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).value = card.get_name()
                row += 1
    
    def task_description(self, col_num):
        self.__set_item_name_cell(col_num, 40, '説明')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).value = board.get_description()
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).value = card.get_desc()
                row += 1

    def task_start_date(self, col_num):
        self.__set_item_name_cell(col_num, 30, '開始日')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).value = card.get_date()
                row += 1

    def task_last_activity_date(self, col_num):
        self.__set_item_name_cell(col_num, 30, '更新日')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).value = card.get_date_last_activity()
                row += 1

    def task_list_name(self, col_num):
        self.__set_item_name_cell(col_num, 20, 'ステータス')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).value = card.get_listname()
                row += 1

    def task_members(self, col_num):
        self.__set_item_name_cell(col_num, 20, '担当者')
        row = DATA_ROW_START
        for board in self.__boards:
            self.__ws.cell(row=row, column=col_num).fill = PHASE_FILL
            self.__ws.cell(row=row, column=col_num).border = BORDER 
            row += 1
            for card in board.get_cards():
                self.__ws.cell(row=row, column=col_num).value = ','.join(card.get_membernames())
                row += 1


    def exportAsExcel(self):
        if len(self.__boards) < 1:
            return

        self.task_ids(1) 
        self.task_names(2)
        self.task_description(3)
        self.task_start_date(4)
        self.task_last_activity_date(5)
        self.task_list_name(6)
        self.task_members(7)
        self.performance(self.__ws.max_column)

        self.__wb.save('test.xlsx')
